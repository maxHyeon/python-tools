from datetime import date
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
    Series
)
# 초기화
from openpyxl.chart.axis import DateAxis
from openpyxl import load_workbook
wb = load_workbook(filename = 'D:\\autoReport\\PER_PerfDayList_excel.xlsx')
wb2 = load_workbook(filename = 'D:\\autoReport\\patchSchedule.xlsx')
sheet_ranges = wb['sheet1']
sheet_ranges2 = wb['Windows']
ws = wb.active
ws2 = wb2.active
def drawChartCPU(stIdx,endIdx) :
	c1 = LineChart()
	c1.title = sheet_ranges['A'+str(stIdx)].value[4:]+'_CPU'
	c1.style = 12	
	c1.x_axis.number_format = 'yyyymmdd'
	c1.x_axis.majorTimeUnit = "days"
	c1.legend.position = 'b'
	data = Reference(ws, min_col=3, min_row=stIdx,  max_row=endIdx)
	series = Series(data, title="CPU BUSY MAX(%)")
	c1.append(series)
	data = Reference(ws, min_col=4, min_row=stIdx,  max_row=endIdx)	
	series = Series(data, title="CPU BUSY AVG(%)")
	c1.append(series)
	dates = Reference(ws, min_col=2, min_row=stIdx, max_row=endIdx)
	c1.set_categories(dates)
	s1 = c1.series[0]
	s1.graphicalProperties.line.solidFill ="5698d4"
	s2 = c1.series[1]
	s2.graphicalProperties.line.solidFill ="ed7d30"
	ws.add_chart(c1, "K"+str(stIdx))
	
def drawChartMem(stIdx,endIdx) :
	c2 = LineChart()
	c2.title = sheet_ranges['A'+str(stIdx)].value[4:]+'_MEM'
	c2.style = 12
	c2.x_axis.number_format ='yyyymmdd'
	c2.x_axis.majorTimeUnit = "days"
	c2.legend.position = 'b'
	data = Reference(ws, min_col=5, min_row=stIdx,  max_row=endIdx)
	series = Series(data, title="MEMORY PCTUSED MAX(%)")
	c2.append(series)
	data = Reference(ws, min_col=6, min_row=stIdx,  max_row=endIdx)	
	series = Series(data, title="MEMORY PCTUSED AVG(%)")
	c2.append(series)
	data = Reference(ws, min_col=7, min_row=stIdx,  max_row=endIdx)
	series = Series(data, title="MEMORY_SWAPPCTUSED_MAX(%)")
	c2.append(series)
	data = Reference(ws, min_col=8, min_row=stIdx,  max_row=endIdx)	
	series = Series(data, title="MEMORY_SWAPPCTUSED_AVG(%)")
	c2.append(series)	
	dates = Reference(ws, min_col=2, min_row=stIdx, max_row=endIdx)
	c2.set_categories(dates)
	s1 = c2.series[0]
	s1.graphicalProperties.line.solidFill ="5698d4"
	s2 = c2.series[1]
	s2.graphicalProperties.line.solidFill ="ed7d30"
	s3 = c2.series[2]
	s3.graphicalProperties.line.solidFill ="adadad"
	s4 = c2.series[3]
	s4.graphicalProperties.line.solidFill ="ffc100"
	ws.add_chart(c2, "P"+str(stIdx))
	
i=2
endHostIdx=2;
startHostIdx=2;

while sheet_ranges['A'+str(i)].value != None:
	if sheet_ranges['A'+str(i)].value == sheet_ranges['A'+str(i+1)].value :
		startHostIdx = endHostIdx;
		print(startHostIdx)
	if sheet_ranges['A'+str(i)].value != sheet_ranges['A'+str(i+1)].value :
		endHostIdx = i+1
		print(endHostIdx)
		drawChartCPU(startHostIdx,endHostIdx)
		drawChartMem(startHostIdx,endHostIdx)
	
	i=i+1


wb.save("D:\\autoReport\\PER_PerfDayList_excel.xlsx")