from datetime import date
from openpyxl import Workbook
import re
# 초기화
from openpyxl.chart.axis import DateAxis
from openpyxl import load_workbook

wb = load_workbook(filename = "copyPaste.xlsx")
sheet_ranges = wb['list']
wb2 = load_workbook(filename = "test2.xlsx")
sheet_ranges2 = wb2['@']
ws = wb.active
ws2 = wb2.active

detailDic = {"\IBM-PeriodicRunner":"IBM BigFix 스케줄러" , "\IBM-HDD_Size": "HDD SIZE 정보 수집 스케줄러", "\IBM-EventLogGather" : "이벤트 로그 백업 스케줄러", "\IBM-DailyCheck": "일일 점검 스케줄러","\EventLogGather":"이벤트 로그 백업 스케줄러" ,"\PeriodicRunner":"IBM BigFix 스케줄러","\EventlogGather":"이벤트 로그 백업 스케줄러","\DailyCheck":"일일 점검 스케줄러","\ITMon_Stop_Start":"티볼리 모니터링 스케줄러"} 

def checkHostIDX(tempHostNm):    
    for row in ws.iter_rows(min_row=2):
        if row[1].value == tempHostNm:
            return row[1].row
            break

def isInData(orgIdx):
    return sheet_ranges['E'+str(orgIdx)].value != None

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    for i in range(startRow,endRow + 1,1):
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        rangeSelected.append(rowSelected)
    return rangeSelected

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):	
        countCol = 0
        for j in range(startCol,endCol+1,1):
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

def pastePrevRow(copyIdx):
    copyData = copyRange(1,copyIdx,4,copyIdx,sheet_ranges)
    #print(copyData)
    ws.insert_rows(copyIdx)
    #print('insert');
    pasteRange(1,copyIdx ,4,copyIdx,sheet_ranges,copyData)

def copyPasteScheduleData(scheduleIdx,pasteIdx):
    sheet_ranges.cell(row = pasteIdx, column = 5).value = str(str(sheet_ranges2.cell(row = scheduleIdx, column = 15).value))
    #print(sheet_ranges.cell(row = pasteIdx, column = 5).value)
    sheet_ranges.cell(row = pasteIdx, column = 6).value = str(str(sheet_ranges2.cell(row = scheduleIdx, column = 23).value)) + ' ' + str(sheet_ranges2.cell(row = scheduleIdx, column = 24).value) + ' ' +str(sheet_ranges2.cell(row = scheduleIdx, column = 25).value) +' ' +str(sheet_ranges2.cell(row = scheduleIdx, column = 26).value) +' ' + str(sheet_ranges2.cell(row = scheduleIdx, column = 27).value) + ' ' +str(sheet_ranges2.cell(row = scheduleIdx, column = 28).value)
    #print(sheet_ranges.cell(row = pasteIdx, column = 6).value)
    sheet_ranges.cell(row = pasteIdx, column = 7).value = str(str(sheet_ranges2.cell(row = scheduleIdx, column = 2).value))
    #print(sheet_ranges.cell(row = pasteIdx, column = 7).value)
    sheet_ranges.cell(row = pasteIdx, column = 8).value = str(str(sheet_ranges2.cell(row = scheduleIdx, column = 9).value)) 
    #print(sheet_ranges.cell(row = pasteIdx, column = 8).value)
    sheet_ranges.cell(row = pasteIdx, column = 9).value = str(detailDic.get(sheet_ranges.cell(row = pasteIdx, column = 7).value,""))
    #print(str(detailDic.get(sheet_ranges.cell(row = pasteIdx, column = 7).value,"")))


for scheRow in ws2.iter_rows(min_row=1):
    if scheRow[0].value == None:
        continue
    #print (scheRow[0].value)
    if checkHostIDX(scheRow[0].value) != None:
        #print('in if');
        tempIdx = checkHostIDX(scheRow[0].value)
        #print (tempIdx)
        if isInData(tempIdx):     
            	
            pastePrevRow(tempIdx)
            copyPasteScheduleData(scheRow[0].row,tempIdx)
            #print(tempIdx)			
        else :        
            #print('not in data');
            copyPasteScheduleData(scheRow[0].row,tempIdx)
            #print(tempIdx)
		
wb.save("shceduleMiddle.xlsx")