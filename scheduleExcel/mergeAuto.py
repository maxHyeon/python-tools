from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl import load_workbook

wb = load_workbook(filename = "shceduleMiddle.xlsx")
sheet_ranges = wb['list']
ws = wb.active

key_column = 2
merge_columns = [1,2,3,4]
start_row = 2
max_row = ws.max_row
key = None

# Iterate all rows in `key_colum`
for row, row_cells in enumerate(ws.iter_rows(min_col=key_column, min_row=start_row,
                                             max_col=key_column, max_row=max_row),
                                start_row):
    if key != row_cells[0].value or row == max_row:
        # moved line below this if
        # if row == max_row: row += 1 
        if not key is None:
            for merge_column in merge_columns:
                ws.merge_cells( start_row=start_row, start_column=merge_column,
                                end_row=row - 1, end_column=merge_column)

                ws.cell(row=start_row, column=merge_column).\
                    alignment = Alignment(horizontal='center', vertical='center')

            start_row = row

        key = row_cells[0].value
    #moved below line here as it was merging last two rows content even if the values differ.
    if row == max_row: row += 1 

wb.save("emc_pool.xlsx")