from datetime import date
from openpyxl import Workbook
import re
# 초기화
from openpyxl.chart.axis import DateAxis
from openpyxl import load_workbook
wb = load_workbook(filename = "test.xlsx")
sheet_ranges = wb['@']

ws = wb.active

delCnt = 0
orgIdx = 1

delTuple = ("Microsoft Corporation","만든 이","Microsoft","N/A","Author","Microsoft Corporation.","$(@%SystemRoot%\system32\invagent.dll,-701)","$(@%SystemRoot%\system32\ErrorDetailsUpdate.dll,-600)","$(@%systemroot%\system32\ErrorDetailsUpdate.dll,-600)","운영 체제의 주 버전 번호입니다.","The major version number of the operating system.","$(@%SystemRoot%\system32\tzsyncres.dll,-101)","$(@%SystemRoot%\system32\spaceman.exe,-2)","Microsoft VisualStudio","$(@%ProgramFiles%\Common Files\Microsoft Shared\OfficeSoftwareProtectionPlatform\osppc.dll,-200)","$(@%systemroot%\system32\osppc.dll,-200)")
delTuple2 = ("\Microsoft\Windows\Backup\Microsoft-Windows-WindowsBackup","C:\Windows\system32\msfeedssync.exe sync","%windir%\system32\tzsync.exe","\Microsoft\Windows\Time Zone\SynchronizeTimeZone","\OfficeSoftwareProtectionPlatform\SvcRestartTask","C:\Windows\System32\msfeedssyncc.exe sync","%systemroot%\system32\sc.exe start osppsvc","C:\Windows\system32\msfeedssyncc.exe sync")

def delRowCheck (idx):
    if sheet_ranges['H'+str(idx)].value in delTuple:
        #print("this is ms")
        return 1
    else :
        #print(sheet_ranges['H'+str(idx)].value)
        return 0

def delRow(idx):
    ws.delete_rows(idx,1)

def delRowData(idx) : 
    sheet_ranges.cell(row = idx, column = 1).value = None
    sheet_ranges.cell(row = idx, column = 2).value = None
    sheet_ranges.cell(row = idx, column = 3).value = None
    sheet_ranges.cell(row = idx, column = 4).value = None
    sheet_ranges.cell(row = idx, column = 5).value = None
    sheet_ranges.cell(row = idx, column = 6).value = None
    sheet_ranges.cell(row = idx, column = 7).value = None
    sheet_ranges.cell(row = idx, column = 8).value = None
    sheet_ranges.cell(row = idx, column = 9).value = None
    sheet_ranges.cell(row = idx, column = 10).value = None
    sheet_ranges.cell(row = idx, column = 11).value = None
    sheet_ranges.cell(row = idx, column = 12).value = None
    sheet_ranges.cell(row = idx, column = 13).value = None
    sheet_ranges.cell(row = idx, column = 14).value = None


for row in ws.iter_rows(min_row=1):
    if row[7].value in delTuple or row[1].value in delTuple2 or row[8].value in delTuple2:
        print(row[7].value)
        delRowData(row[1].row)

#while sheet_ranges['H'+str(orgIdx)].value != None:
#    if delRowCheck(orgIdx) == 1:
#        delCnt = delCnt + 1
#        delRow(orgIdx)
#        #print("delCnt = "+ str(delCnt) + " orgIdx ="+str(orgIdx) )
#    else:
#        orgIdx = orgIdx + 1
#        #print("delCnt = "+ str(delCnt) + " orgIdx ="+str(orgIdx) )

wb.save("test2.xlsx")